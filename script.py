import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, borders, numbers
import re
import os

def main_year(base_path: str):
    purchases_and_expenses_row = 37
    purchases_and_expenses_col = 0
    year = parse_year_from_path(base_path)
    if year == None:
        return
    leap_year = year % 4 == 0 and (year % 100 != 0 or year % 400 == 0)
    months = {
        '01': { 'name':  "January"    ,'present': False, 'path':None  ,'days': 31, },
        '02': { 'name':  "February"   ,'present': False, 'path':None  ,'days': 29 if leap_year else 28, },
        '03': { 'name':  "March"      ,'present': False, 'path':None  ,'days': 31, },
        '04': { 'name':  "April"      ,'present': False, 'path':None  ,'days': 30, },
        '05': { 'name':  "May"        ,'present': False, 'path':None  ,'days': 31, },
        '06': { 'name':  "June"       ,'present': False, 'path':None  ,'days': 30, },
        '07': { 'name':  "July"       ,'present': False, 'path':None  ,'days': 31, },
        '08': { 'name':  "August"     ,'present': False, 'path':None  ,'days': 31, },
        '09': { 'name':  "September"  ,'present': False, 'path':None  ,'days': 30, },
        '10': { 'name':  "October"    ,'present': False, 'path':None  ,'days': 31, },
        '11': { 'name':  "November"   ,'present': False, 'path':None  ,'days': 30, },
        '12': { 'name':  "December"   ,'present': False, 'path':None  ,'days': 31, },
    }
    find_months(months, base_path)
    year_file_path = "./Fremont Book - {}.xlsx".format(year)
    if not os.path.exists(year_file_path):
        yearbook = create_workbook(year_file_path)
    else:
        os.remove(year_file_path)
        yearbook = create_workbook(year_file_path)

    for m in months:
        if months[m]['present'] == False:
            continue
        month_file = open_workbook(months[m]['path'])
        yearbook.create_sheet(months[m]['name'])
        month_sheet = yearbook[months[m]['name']]
        month_sheet.append([
            "DATE",
            "CHECK#",
            "INVOICE#",
            "VENDOR NAME",
            "",
            "",
            "CODE",
            "COST",
            "CASH PAID OUT",
            ""
        ])
        month_count = 0
        for day in range(1, months[m]['days'] + 1):
            day_sheet = month_file[day.__str__()]
            day_data = []
            day_start = True
            
            if (year == 2023) or (year == 2022 and int(m) >= 7):
                purchases_and_expenses_row = 38
                purchases_and_expenses_col = 2
            """ sanity_check_row = day_sheet[purchases_and_expenses_row - 1] """
            """ i = 10 """
            """ print("{} - {} - {} --- ".format(year, months[m]['name'], day), end="") """
            """ for c in sanity_check_row: """
            """     if i == 0: """
            """         break """
            """     i = i - 1 """
            """     print(c.value, end=", ") """
            """ print("") """
            
            for rows in day_sheet.iter_rows(
                min_row=purchases_and_expenses_row, 
                min_col=purchases_and_expenses_col,
                max_col=purchases_and_expenses_col + 9
            ):
                row_tmp = []

                row_tmp = date_column(row_tmp, day_start, (m, day, year))
                day_start = False

                empty_row = True
                for cell in rows:
                    if cell.internal_value != None and cell.internal_value != 0 and not_empty_string(cell.internal_value):
                        empty_row = False
                    if type(cell.internal_value) == str:
                        row_tmp.append(cell.internal_value.strip())
                    else:
                        row_tmp.append(cell.internal_value)

                if empty_row == False:
                    day_data.append(row_tmp)
                    month_count = month_count + 1
                else:
                    break

            for d in day_data:
                month_sheet.append(d)
        month_count = month_count + 1
        month_sheet.append(["", "", "", "", "", "", "Totals", "=sum(H2:H{})".format(month_count), "", "=sum(J2:J{})".format(month_count)])
        format_sheet(yearbook, months[m]['name'], month_count)
    yearbook.remove(yearbook['Sheet'])
    yearbook.save(year_file_path)

def format_sheet(book: openpyxl.Workbook, sheet_name: str, rows: int):
    sheet = book[sheet_name]
    font = Font(bold=True)
    color = 'FFFFCC99'
    border_style = Side(border_style='thin', color='FF000000')
    border = Border(outline=True, left=border_style, right=border_style, top=border_style, bottom=border_style)
    fill = PatternFill(fill_type='solid',start_color=color)
    for r in sheet.iter_rows(0,1,0,10):
        for c in r:
            c.font = font
    for c in sheet.iter_cols(10,10):
        for r in c:
            r.fill = fill
    c = sheet['I1']
    c.fill = fill
    c = sheet['H1']
    c.fill = fill
    for c in sheet.iter_cols(8,8):
        for r in c:
            r.fill = fill
    for c in sheet.iter_cols(0, 10, 0, rows):
        for r in c:
            r.border = border

    c = sheet["H{}".format(rows + 1)]
    c.border = border
    c.number_format = numbers.FORMAT_CURRENCY_USD
    c.font = font
    c = sheet["J{}".format(rows + 1)]
    c.border = border
    c.number_format = numbers.FORMAT_CURRENCY_USD
    c.font = font

    c = sheet["G{}".format(rows + 1)]
    c.font = font

    border_style = Side(border_style='thin', color='FF000000')
    border_style_2 = Side(border_style='none')
    border = Border(outline=True, left=border_style_2, right=border_style_2, top=border_style, bottom=border_style)
    for c in sheet.iter_cols(4,6,0, rows):
        for r in c:
            r.border = border

    sheet.column_dimensions['A'].width = 10
    sheet.column_dimensions['B'].width = 8.14
    sheet.column_dimensions['C'].width = 10
    sheet.column_dimensions['D'].width = 10
    sheet.column_dimensions['E'].width = 10
    sheet.column_dimensions['F'].width = 10
    sheet.column_dimensions['G'].width = 10
    sheet.column_dimensions['H'].width = 10
    sheet.column_dimensions['I'].width = 10
    sheet.column_dimensions['J'].width = 10
    sheet.merge_cells('D1:F1')
    sheet.merge_cells('I1:J1')


def date_column(row: list, first_row_of_day: bool, date: tuple):
    m, d, y = date
    if d < 10:
        d = "0{}".format(d)
    else:
        d = "{}".format(d)
    if first_row_of_day:
        row.append("{}/{}/{}".format(m,d,y))
        first_row_of_day = False
    else:
        row.append(None)
    return row


def create_workbook(name: str):
    new_workbook = openpyxl.Workbook()
    new_workbook.save(name)
    return new_workbook

def open_workbook(name: str):
    if name.find(".xlsx") == -1:
        name = name + ".xlsx"
    workbook = openpyxl.load_workbook(name)
    return workbook

def parse_year_from_path(path: str):
    pattern = r'\b20(1\d|2\d)\b'
    match = re.search(pattern, path)
    if match:
        return int(match.group())
    else:
        return None

def parse_month_from_path(path: str):
    month_codes = {'jan': '01', 'feb': '02', 'mar': '03', 'apr': '04',
                   'may': '05', 'jun': '06', 'jul': '07', 'aug': '08',
                   'sep': '09', 'oct': '10', 'nov': '11', 'dec': '12'}
    pattern = '|'.join(month_codes.keys())
    match = re.search(pattern, path, re.IGNORECASE)
    if match:
        return month_codes[match.group().lower()]
    else:
        return None 

def find_months(months: dict, base_path: str):
    for f in os.listdir(base_path):
        m = parse_month_from_path(f)
        if m != None:
            months[m]['present'] = True
            months[m]['path'] = "{}/{}".format(base_path, f)

def not_empty_string(cell_val):
    if type(cell_val) == str:
        if len(cell_val.strip()) == 0:
            return False
    return True

def main():
    for f in os.listdir('books'):
        main_year("./books/{}".format(f))

main()
